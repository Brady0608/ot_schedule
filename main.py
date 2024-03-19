import openpyxl
import pandas as pd
import datetime
import time
from openpyxl.styles import Font
import re
from pprint import pprint

therapist_list = []
schedule_time = []
schedule_week = []
inpatient_dict = {}
OPD_dict = {}
therapist_schedule = {}
week = ['1', '2', '3', '4', '5']
input_day = input('請輸入周一日期，格式為XXXX/XX/XX:')
t = time.strptime(input_day, "%Y/%m/%d")
y, m, d = t[0:3]
print(datetime.date(y, m, d) + datetime.timedelta(days=1))
print(str((datetime.date(y, m, d))))
new_day = (datetime.date(y, m, d))
date_5_days = [(datetime.date(y, m, d))]
date_5_days_o =[]
for i in range(4):  # 列出為未來五天日期
    new_day = new_day + datetime.timedelta(days=1)
    date_5_days.append(new_day)

for i in range(5):  # 日期西元轉為民國
    date_5_days[i] = str(date_5_days[i])
    # s_t =
    date_5_days[i] = date_5_days[i].replace(date_5_days[i][:4], str(int(date_5_days[i][:4]) - 1911))
    date_5_days_o.append(date_5_days[i])
    date_5_days[i] = date_5_days[i].replace('-', '/')
# print(date_5_days)
Week = ['   (一)', '   (二)', '   (三)', '   (四)', '   (五)']
col_week = []
for i in range(len(date_5_days)):  # excel日期值
    col_week.append(date_5_days[i] + Week[i])
print(col_week)

excel_file = './病人清單.xlsx'
wb = openpyxl.load_workbook(filename=excel_file)
sheet = wb['Inpatient']

Inpatient_data = pd.read_excel(excel_file, sheet_name='Inpatient')  # 讀取SHEET
OPD_data = pd.read_excel(excel_file, sheet_name='OPD')
# print(Inpatient_data[Inpatient_data['治療師']=='陳怡儒'])
therapist_list = sorted(list(set(list(Inpatient_data['治療師']))))  # 取得治療師名單
# print(therapist_list)
schedule_time = sorted(list(set(list(Inpatient_data['排程時間']))))  # 取得排程時間，並做排序
# print(schedule_time)
# datetime.time.strftime()
for therapist in therapist_list:
    inpatient_dict[therapist] = Inpatient_data.loc[
        Inpatient_data['治療師'] == therapist, ['姓名', '排程時間']].values.tolist()  # 取得每位治療師各自病人名單並以list格式存在字典裡
    OPD_dict[therapist] = OPD_data.loc[
        OPD_data['治療師'] == therapist, ['姓名', '排程時間', '排程天數']].values.tolist()  # 取得每位治療師各自病人名單並以list格式存在字典裡
pprint(inpatient_dict)

for therapist in therapist_list:
    for ot in OPD_dict[therapist]:
        ot.append(list(str(ot[2])))
        del ot[2]  # 將OPD排程天數轉為LIST存在字典裡

    for ot in inpatient_dict[therapist]:  # 將InP新增排程天數轉為LIST存在字典裡
        ot.append(['1', '2', '3', '4', '5'])

all_patient_dict = {therapist: inpatient_dict[therapist] + OPD_dict[therapist] for therapist in
                    therapist_list}  # 將兩字典合併
# print(all_patient_dict)
therapist_schedule_dict = {}
week_total_dict = {}

for therapist in therapist_list:
    week_total = 0
    # f = {}
    # for n_week in week:
    #     f[n_week] = None
    #     for t_time in schedule_time:
    #         f[n_week][t_time] = None
    # 每位治療師病人行程表,字典預設值
    f = {'1': {datetime.time(8, 30): None, datetime.time(9, 45): None, datetime.time(11, 0): None,
               datetime.time(13, 30): None, datetime.time(14, 45): None, datetime.time(16, 0): None},
         '2': {datetime.time(8, 30): None, datetime.time(9, 45): None, datetime.time(11, 0): None,
               datetime.time(13, 30): None, datetime.time(14, 45): None, datetime.time(16, 0): None},
         '3': {datetime.time(8, 30): None, datetime.time(9, 45): None, datetime.time(11, 0): None,
               datetime.time(13, 30): None, datetime.time(14, 45): None, datetime.time(16, 0): None},
         '4': {datetime.time(8, 30): None, datetime.time(9, 45): None, datetime.time(11, 0): None,
               datetime.time(13, 30): None, datetime.time(14, 45): None, datetime.time(16, 0): None},
         '5': {datetime.time(8, 30): None, datetime.time(9, 45): None, datetime.time(11, 0): None,
               datetime.time(13, 30): None, datetime.time(14, 45): None, datetime.time(16, 0): None}}

    for patient_num in range(len(all_patient_dict[therapist])):
        # print(len(all_patient_dict[therapist]))
        if bool(re.search(r'\d', all_patient_dict[therapist][patient_num][0])):  # 住院病人若有出院,排程天數修改
            out_num = int(all_patient_dict[therapist][patient_num][0][-1])  # out_num為出院周幾
            print(all_patient_dict[therapist][patient_num][0])
            print(all_patient_dict[therapist][patient_num][2][:out_num - 1])
            all_patient_dict[therapist][patient_num][0] += " "
            if out_num <= 5:
                all_patient_dict[therapist][patient_num][2] = all_patient_dict[therapist][patient_num][2][:out_num - 1]

        for day in week:
            if day in all_patient_dict[therapist][patient_num][2]:
                for t_time in schedule_time:
                    if all_patient_dict[therapist][patient_num][1] == t_time:
                        if f[day][t_time] == None:
                            f[day][t_time] = [all_patient_dict[therapist][patient_num][0]]
                            week_total += 1
                        else:
                            f[day][t_time].append(all_patient_dict[therapist][patient_num][0])
                            week_total += 1

    therapist_schedule_dict[therapist] = f
    week_total_dict[therapist] = week_total

pprint(week_total_dict)
# pprint.pprint(therapist_schedule_dict)
df_dict = {}
characters = "[],'"
col = ['B', 'C', 'D', 'E', 'F']
counts = {'1': {datetime.time(8, 30): 0, datetime.time(9, 45): 0, datetime.time(11, 0): 0,
                datetime.time(13, 30): 0, datetime.time(14, 45): 0, datetime.time(16, 0): 0},
          '2': {datetime.time(8, 30): 0, datetime.time(9, 45): 0, datetime.time(11, 0): 0,
                datetime.time(13, 30): 0, datetime.time(14, 45): 0, datetime.time(16, 0): 0},
          '3': {datetime.time(8, 30): 0, datetime.time(9, 45): 0, datetime.time(11, 0): 0,
                datetime.time(13, 30): 0, datetime.time(14, 45): 0, datetime.time(16, 0): 0},
          '4': {datetime.time(8, 30): 0, datetime.time(9, 45): 0, datetime.time(11, 0): 0,
                datetime.time(13, 30): 0, datetime.time(14, 45): 0, datetime.time(16, 0): 0},
          '5': {datetime.time(8, 30): 0, datetime.time(9, 45): 0, datetime.time(11, 0): 0,
                datetime.time(13, 30): 0, datetime.time(14, 45): 0, datetime.time(16, 0): 0}}

for therapist in therapist_list:
    df = pd.DataFrame.from_dict(therapist_schedule_dict[therapist])

    df.dropna(how='all', inplace=True)
    # pprint(therapist)
    # pprint(df)
    for Week in df.columns:
        for t_time in schedule_time:
            try:
                if df[Week][t_time]:

                    for i in range(len(df[Week][t_time])):
                        counts[Week][t_time] += 1
                        if len(df[Week][t_time][i]) > 3:
                            if df[Week][t_time][i][-1] != " ":
                                df[Week][t_time][i] = df[Week][t_time][i][-3:]
                            else:
                                df[Week][t_time][i] = df[Week][t_time][i][-4:]
                        if len(df[Week][t_time][i]) == 2:
                            df[Week][t_time][i] = df[Week][t_time][i][:1] + "  " + df[Week][t_time][i][1:]

                        df[Week][t_time][i] = "(  )" + df[Week][t_time][i] + " "
                    df[Week][t_time] = ''.join(x for x in str(df[Week][t_time]) if x not in characters)
                    #             break
                    # df[Week][t_time].to_string()
                    # pprint.pprint(df[Week][t_time])
            except KeyError:
                # print("Got KeyError")
                continue
    df.columns = col_week
    df_dict[therapist] = df

pprint(counts)
font_ = Font(
    name="標楷體",
    size=14,
    # italic=True,
    # color='ffff00',
    bold=False,
    strike=None
)

font_1 = Font(
    name="標楷體",
    size=14,
    # italic=True,
    # color='ffff00',
    bold=True,
    strike=None
)

font_2 = Font(
    name="標楷體",
    size=12,
    # italic=True,
    # color='ffff00',
    bold=False,
    strike=None
)
with pd.ExcelWriter(date_5_days_o[0] + '~' + date_5_days_o[4] + '各時段人次統計.xlsx', engine='openpyxl') as writer:
    df_counts = pd.DataFrame.from_dict(counts)
    df_counts.columns = col_week
    df_counts.to_excel(writer, sheet_name='各時段人次統計', startrow=1)
    ws = writer.sheets['各時段人次統計']
    for x in col:
        ws.column_dimensions[x].width = 32.5
    ws.column_dimensions['A'].width = 32.5
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = openpyxl.styles.Alignment(wrap_text=True, vertical='center',horizontal='center')
            cell.font = font_
    ws.print_options.gridLines = True  # 列印格線
    ws.print_options.horizontalCentered = True  # 列印置中
    ws.page_setup.scale = 71  # 列印縮放比例
    ws.set_printer_settings(ws.PAPERSIZE_A4, ws.ORIENTATION_LANDSCAPE) # A4 橫式列印
    ws.merge_cells(start_row=1, start_column=3, end_row=1, end_column=5)
    ws['C1'] = col_week[0] +' ~ ' + col_week[4] + '各時段人次統計'


with pd.ExcelWriter(date_5_days_o[0] + '~' + date_5_days_o[4] + "病患出席紀錄表.xlsx", engine='openpyxl') as writer:
    # df_counts = pd.DataFrame.from_dict(counts)
    # df_counts.columns = col_week
    # df_counts.to_excel(writer, sheet_name='各時段人次統計', startrow=1)
    # ws = writer.sheets['各時段人次統計']
    # for x in col:
    #     ws.column_dimensions[x].width = 32.5
    # ws.column_dimensions['A'].width = 32.5
    # for row in ws.iter_rows():
    #     for cell in row:
    #         cell.alignment = openpyxl.styles.Alignment(wrap_text=True, vertical='center',horizontal='center')
    #         cell.font = font_
    # ws.print_options.gridLines = True  # 列印格線
    # ws.print_options.horizontalCentered = True  # 列印置中
    # ws.page_setup.scale = 71  # 列印縮放比例
    # ws.set_printer_settings(ws.PAPERSIZE_A4, ws.ORIENTATION_LANDSCAPE) # A4 橫式列印
    # ws.merge_cells(start_row=1, start_column=3, end_row=1, end_column=5)
    # ws['C1'] = col_week[0] +' ~ ' + col_week[4] + '各時段人次統計'
    for therapist in therapist_list:
        df_dict[therapist].to_excel(writer, sheet_name=therapist, startrow=1)
        worksheet = writer.sheets[therapist]
        worksheet.merge_cells(start_row=1, start_column=3, end_row=1, end_column=5)
        worksheet['B2'].font = worksheet['C2'].font = worksheet['D2'].font = worksheet['E2'].font = worksheet[
            'F2'].font = font_1
        worksheet['B2'].alignment = worksheet['C2'].alignment = worksheet['D2'].alignment = worksheet['E2'].alignment = \
            worksheet[
                'F2'].alignment = openpyxl.styles.Alignment(wrap_text=True, vertical='center', horizontal='right')
        worksheet['C1'] = therapist
        worksheet['C1'].font = font_
        worksheet['C1'].alignment = openpyxl.styles.Alignment(wrap_text=True, horizontal='center')

        worksheet['F1'] = '本周總人次: ' + str(week_total_dict[therapist])
        worksheet['F1'].font = font_
        worksheet['F1'].alignment = openpyxl.styles.Alignment(wrap_text=True, horizontal='center')

        for row in worksheet.iter_rows(min_row=3, min_col=2):
            for cell in row:
                cell.alignment = openpyxl.styles.Alignment(wrap_text=True, vertical='top')
                cell.font = font_
        for x in col:
            worksheet.column_dimensions[x].width = 32.5
        for i in [3, 4, 5]:
            worksheet.row_dimensions[i].height = 160
        for i in [2, 6, 7, 8]:
            worksheet.row_dimensions[i].height = 25
        worksheet.row_dimensions[1].height = 20

        worksheet['B6'] = worksheet['C6'] = worksheet['D6'] = worksheet['E6'] = worksheet[
            'F6'] = '門診(  ) + 住院(  ) =     '
        for row in worksheet.iter_rows(min_row=6, min_col=1, max_row=9, max_col=7):
            for cell in row:
                cell.alignment = openpyxl.styles.Alignment(wrap_text=True, horizontal='left', vertical='center')
                cell.font = font_2
        worksheet['A7'] = 'NEW'
        worksheet['A8'] = 'DC'
        worksheet['A7'].alignment = worksheet['A8'].alignment = openpyxl.styles.Alignment(wrap_text=True,
                                                                                          horizontal='center',
                                                                                          vertical='center')
        worksheet.set_printer_settings(worksheet.PAPERSIZE_A4, worksheet.ORIENTATION_LANDSCAPE)  # 列印設定
        # worksheet.print_area ='A1:F8'
        worksheet.print_options.gridLines = True  # 列印格線
        worksheet.print_options.horizontalCentered = True  # 列印置中
        worksheet.print_options.verticalCentered = True
        worksheet.page_setup.scale = 78  # 列印縮放比例
